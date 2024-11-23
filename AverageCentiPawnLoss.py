import chess
from stockfish import Stockfish
from openpyxl import Workbook
import random
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import os

# Function to check if the game is over
def isOurGameOver(board):
    return (
        board.is_checkmate() or 
        board.is_stalemate() or 
        board.is_insufficient_material() or 
        board.is_seventyfive_moves() or 
        board.is_fivefold_repetition()
    )


# Initialize Stockfish with the path hidden and engine parameters
stockfish = Stockfish("./stockfish-windows-x86-64-avx2.exe")  # Path to Stockfish engine
stockfish.set_depth(10)
stockfish.set_skill_level(20)

# Update engine parameters (tuning Stockfish's behavior)
stockfish.update_engine_parameters({
    "Debug Log File": "",
    "Contempt": 24,  # Slightly favors strong moves over overly defensive ones
    "Min Split Depth": 0,
    "Threads": 4,  # Adjust based on your system's logical processors
    "Ponder": "true",  # Enables pondering to calculate during opponent's turn
    "Hash": 8192,  # Set to 8 GB for systems with 16 GB RAM; adjust for your system
    "MultiPV": 3,  # 3 best move for maximum strength
    "Skill Level": 20,  # Maximum skill level
    "Move Overhead": 30,  # Additional time for move generation
    "Minimum Thinking Time": 50,  # Ensure enough time for deeper calculations
    "Slow Mover": 100,  # Default speed
    "UCI_Chess960": "false",
    "UCI_LimitStrength": "false",  # No limit on engine strength
})

# Create a new Excel workbook and sheets for game and turn data
wb = Workbook()
ws_game = wb.active
ws_game.title = "Chess Games"
ws_turn = wb.create_sheet("Game And Turn")

# Adding headers for game data table (Game, Ply number, Winner)
ws_game.cell(row=1, column=1).value = "Game Number"
ws_game.cell(row=1, column=2).value = "Ply Number"
ws_game.cell(row=1, column=3).value = "Who Won"

# Adding headers for turn data
ws_turn.cell(row=1, column=1).value = "Number"

# Initialize counters
game_number = 1
row = 2
colForTurn = 2

# Initialize the GUI window
root = tk.Tk()
root.title("Chess Game Status")
root.geometry("600x600")

# Create a canvas to display the chessboard
canvas = tk.Canvas(root, width=480, height=480)
canvas.pack()

# Load piece images
piece_images = {}
piece_names = {
    'K': 'white-king', 'Q': 'white-queen', 'R': 'white-rook', 'B': 'white-bishop', 'N': 'white-knight', 'P': 'white-pawn',
    'k': 'king', 'q': 'queen', 'r': 'rook', 'b': 'bishop', 'n': 'knight', 'p': 'pawn'
}
for piece, filename in piece_names.items():
    image = Image.open(os.path.join("icons", f"{filename}.png"))
    image = image.resize((30, 30))
    piece_images[piece] = ImageTk.PhotoImage(image)

# Function to update the chessboard on the GUI with piece images
def update_chessboard(board):
    canvas.delete("all")  # Clear the canvas
    square_size = 60

    for row in range(8):
        for col in range(8):
            x1 = col * square_size
            y1 = row * square_size
            x2 = x1 + square_size
            y2 = y1 + square_size
            color = "white" if (row + col) % 2 == 0 else "gray"
            canvas.create_rectangle(x1, y1, x2, y2, fill=color)

            square_index = (7 - row) * 8 + col
            piece = board.piece_at(square_index)
            
            if piece:
                piece_symbol = piece.symbol()

                if piece_symbol in piece_images:
                    canvas.create_image(
                        x1 + square_size / 2,
                        y1 + square_size / 2,
                        image=piece_images[piece_symbol]
                    )
                else:
                    canvas.create_text(
                        x1 + square_size / 2,
                        y1 + square_size / 2,
                        text=piece_symbol,
                        fill="red",
                        font=("Helvetica", 20)
                    )



# Function to start a new game
def start_game():
    global game_number, row, colForTurn

    side = True  # True for white, False for black

    while game_number < 10:
        board = chess.Board()
        update_chessboard(board)
        root.update()  # Update the GUI window
    
        # Play game logic here
        while not isOurGameOver(board):
            stockfish.set_fen_position(board.fen())  # Update the board position in Stockfish
            move_list = stockfish.get_top_moves()  # Get possible moves from Stockfish
            try:
                if len(move_list) > 1:
                    
                    centipawn_values = [move["Centipawn"] for move in move_list]

                    median = centipawn_values[len(centipawn_values) // 2]

                    if side:
                        filtered_moves = [move for move in move_list if move["Centipawn"] >= median]
                    else:
                        filtered_moves = [move for move in move_list if move["Centipawn"] <= median]

                    move = random.choice(filtered_moves)["Move"]
                    board.push_san(move)
                else:
                    board.push_san(move_list[0]["Move"])

            except TypeError:
                print("ERROR OCCURRED: Check for side or move issue")

            side = not side  # Switch sides after each move

            update_chessboard(board)
            root.update()  # Update the GUI window

        # Record game result
        ply_number = board.ply()
        ws_game.cell(row=row, column=1).value = f"Game {game_number}"
        ws_game.cell(row=row, column=2).value = ply_number

        if board.outcome().result() == "1-0":
            ws_game.cell(row=row, column=3).value = "White"
            print("White Won")
        elif board.outcome().result() == "0-1":
            ws_game.cell(row=row, column=3).value = "Black"
            print("Black Won")
        else:
            ws_game.cell(row=row, column=3).value = "Draw"
            print("Draw")

        row += 1
        game_number += 1
        colForTurn += 1

    # Save the workbook with the results
    wb.save("chess_games_with_turns.xlsx")
    messagebox.showinfo("Game Over", "Chess games completed and saved!")

# Start the game when the GUI window is ready
start_button = tk.Button(root, text="Start Game", command=start_game,
                         font=("Helvetica", 16, "bold"),  # Font style and size
                         fg="white",  # Text color
                         bg="green",  # Background color
                         relief="raised",  # Raised 3D effect
                         bd=5,  # Border width
                         padx=20, pady=10,  # Horizontal and vertical padding
                         activebackground="darkgreen",  # Color when the button is pressed
                         activeforeground="white")  # Text color when the button is pressed
start_button.pack(pady=20)  # Add vertical space around the button

# Run the tkinter event loop
root.mainloop()